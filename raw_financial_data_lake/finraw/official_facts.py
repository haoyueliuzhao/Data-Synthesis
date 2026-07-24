from __future__ import annotations

import json
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook


OFFICIAL_FACT_SOURCE_IDS = (
    "nbs_official_statistics",
    "pboc_official_statistics",
    "safe_official_statistics",
    "sse_market_statistics",
    "szse_market_statistics",
)


def iter_official_fact_inputs(db, report: dict[str, Any]) -> Iterable[dict[str, Any]]:
    placeholders = ",".join("?" for _ in OFFICIAL_FACT_SOURCE_IDS)
    rows = db.fetchall(
        f"""
        SELECT rr.raw_record_id, rr.raw_object_id, rr.source_id,
               rr.record_key, rr.record_json, rr.entity_hint,
               rr.metric_hint, rr.period_hint, ro.storage_uri,
               ro.source_publish_date, ro.retrieval_time
        FROM raw_records rr
        JOIN raw_objects ro ON ro.raw_object_id = rr.raw_object_id
        WHERE rr.record_type = 'official_publication'
          AND rr.source_id IN ({placeholders})
          AND ro.validation_status = 'passed'
        ORDER BY rr.source_id, rr.record_key
        """,
        OFFICIAL_FACT_SOURCE_IDS,
    )
    for raw_row in rows:
        row = dict(raw_row)
        payload = _json_value(row.get("record_json"))
        storage_uri = row.get("storage_uri") or payload.get("storage_uri")
        path = Path(str(storage_uri)) if storage_uri else None
        if path is None or not path.exists():
            report["skipped_counts"]["official_publication_missing_file"] += 1
            continue
        try:
            if row["source_id"] == "safe_official_statistics":
                yield from _safe_inputs(row, path)
            elif row["source_id"] == "nbs_official_statistics":
                yield from _nbs_inputs(row, path)
            elif row["source_id"] == "pboc_official_statistics":
                yield from _pboc_inputs(row, path)
            elif row["source_id"] == "sse_market_statistics":
                yield from _sse_inputs(row, path)
            elif row["source_id"] == "szse_market_statistics":
                yield from _szse_inputs(row, path)
        except (OSError, ValueError, KeyError, IndexError) as exc:
            report["skipped_counts"]["official_publication_parse_error"] += 1
            report.setdefault("official_parse_errors", []).append(
                {"record_key": row.get("record_key"), "error": str(exc)}
            )


def _safe_inputs(row: dict[str, Any], path: Path) -> Iterable[dict[str, Any]]:
    record_key = str(row.get("record_key") or "")
    if record_key == "safe_bop_bpm6_timeseries":
        yield from _safe_bop(row, path)
    elif record_key == "safe_external_debt_timeseries_2014_plus":
        yield from _safe_external_debt(row, path)
    elif record_key == "safe_official_reserves_2025_xlsx":
        yield from _safe_reserves(row, path)
    elif record_key == "safe_fx_market_turnover_2026_usd":
        yield from _safe_fx_turnover(row, path)


def _safe_bop(row: dict[str, Any], path: Path) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet_name, frequency in (
        ("年度BOP（美元） ", "annual"),
        ("季度BOP（美元）", "quarterly"),
    ):
        sheet = workbook[sheet_name]
        headers = list(next(sheet.iter_rows(min_row=4, max_row=4, values_only=True)))
        values = list(next(sheet.iter_rows(min_row=6, max_row=6, values_only=True)))
        for period, raw_value in zip(headers[1:], values[1:]):
            parsed = _bop_period(period)
            value = _decimal(raw_value)
            if parsed is None or value is None:
                continue
            start, end, year, quarter = parsed
            yield _input(
                row,
                metric_id="current_account_balance_current_usd",
                value=value,
                unit="hundred_million USD",
                currency="USD",
                period_start=start,
                period_end=end,
                fiscal_year=year,
                fiscal_quarter=quarter,
                source_field_name="BPM6:1. Current account",
                frequency=frequency,
                period_scope="calendar_period",
                sheet=sheet_name.strip(),
            )


def _safe_external_debt(row: dict[str, Any], path: Path) -> Iterable[dict[str, Any]]:
    sheet = load_workbook(path, read_only=True, data_only=True).worksheets[0]
    headers = list(next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    target = None
    for values in sheet.iter_rows(min_row=4, values_only=True):
        if str(values[0] or "").strip() == "外债总额头寸":
            target = list(values)
            break
    if target is None:
        raise ValueError("SAFE external debt total row not found")
    for period, raw_value in zip(headers[1:], target[1:]):
        parsed = _cn_quarter_end(period)
        value = _decimal(raw_value)
        if parsed is None or value is None:
            continue
        start, end, year, quarter = parsed
        yield _input(
            row,
            metric_id="external_debt_total_current_usd",
            value=value,
            unit="hundred_million USD",
            currency="USD",
            period_start=start,
            period_end=end,
            fiscal_year=year,
            fiscal_quarter=quarter,
            source_field_name="外债总额头寸",
            frequency="quarterly",
            period_scope="calendar_point_in_time",
            sheet=sheet.title,
        )


def _safe_reserves(row: dict[str, Any], path: Path) -> Iterable[dict[str, Any]]:
    sheet = load_workbook(path, read_only=True, data_only=True).worksheets[0]
    headers = list(next(sheet.iter_rows(min_row=4, max_row=4, values_only=True)))
    metric_rows = {
        8: ("official_foreign_exchange_reserves_current_usd", "1. Foreign currency reserves"),
        19: ("official_reserve_assets_current_usd", "Official reserve assets total"),
    }
    for row_number, (metric_id, source_field) in metric_rows.items():
        values = list(next(sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True)))
        for column in range(1, min(len(headers), len(values)), 2):
            period = headers[column]
            value = _decimal(values[column])
            parsed = _month_period(period)
            if parsed is None or value is None:
                continue
            start, end, year = parsed
            yield _input(
                row,
                metric_id=metric_id,
                value=value,
                unit="hundred_million USD",
                currency="USD",
                period_start=start,
                period_end=end,
                fiscal_year=year,
                fiscal_quarter=None,
                source_field_name=source_field,
                frequency="monthly",
                period_scope="calendar_point_in_time",
                sheet=sheet.title,
            )


def _safe_fx_turnover(row: dict[str, Any], path: Path) -> Iterable[dict[str, Any]]:
    sheet = load_workbook(path, read_only=True, data_only=True).worksheets[0]
    headers = list(next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    values = list(next(sheet.iter_rows(min_row=39, max_row=39, values_only=True)))
    for period, raw_value in zip(headers[1:], values[1:]):
        parsed = _month_period(period)
        value = _decimal(raw_value)
        if parsed is None or value is None:
            continue
        start, end, year = parsed
        yield _input(
            row,
            metric_id="fx_market_turnover_current_usd",
            value=value,
            unit="hundred_million USD",
            currency="USD",
            period_start=start,
            period_end=end,
            fiscal_year=year,
            fiscal_quarter=None,
            source_field_name="五、合计",
            frequency="monthly",
            period_scope="calendar_period",
            sheet=sheet.title,
        )


def _nbs_inputs(row: dict[str, Any], path: Path) -> Iterable[dict[str, Any]]:
    record_key = str(row.get("record_key") or "")
    if record_key == "nbs_gdp_2025_q2":
        table = pd.read_html(str(path))[0]
        values = table.loc[table.iloc[:, 0].astype(str).str.strip() == "GDP"].iloc[0]
        specs = (
            ("2025-04-01", "2025-06-30", "Q2", values.iloc[1], values.iloc[3], "quarterly"),
            ("2025-01-01", "2025-06-30", "H1", values.iloc[2], values.iloc[4], "semiannual_cumulative"),
        )
        for start, end, quarter, level, growth, frequency in specs:
            yield _input(row, metric_id="gdp_current_cny", value=level, unit="hundred_million CNY", currency="CNY", period_start=start, period_end=end, fiscal_year=2025, fiscal_quarter=quarter, source_field_name=f"GDP:{quarter}:absolute", frequency=frequency, period_scope="calendar_period")
            yield _input(row, metric_id="real_gdp_growth_pct", value=growth, unit="percent", currency=None, period_start=start, period_end=end, fiscal_year=2025, fiscal_quarter=quarter, source_field_name=f"GDP:{quarter}:yoy_growth", frequency=frequency, period_scope="calendar_period")
    elif record_key == "nbs_main_indicators_2025":
        table = pd.read_html(str(path))[1]
        values = table.loc[table.iloc[:, 0].astype(str).str.contains("国内生产总值", na=False)].iloc[0]
        yield _input(row, metric_id="gdp_current_cny", value=values.iloc[3], unit="hundred_million CNY", currency="CNY", period_start="2025-01-01", period_end="2025-12-31", fiscal_year=2025, fiscal_quarter="FY", source_field_name="国内生产总值:全年绝对量", frequency="annual", period_scope="calendar_period")
        yield _input(row, metric_id="real_gdp_growth_pct", value=values.iloc[4], unit="percent", currency=None, period_start="2025-01-01", period_end="2025-12-31", fiscal_year=2025, fiscal_quarter="FY", source_field_name="国内生产总值:全年同比增长", frequency="annual", period_scope="calendar_period")
    elif record_key == "nbs_industrial_production_2025_12":
        values = pd.read_html(str(path))[0].iloc[2]
        for start, field, value, frequency in (("2025-12-01", "12月同比增长", values.iloc[2], "monthly"), ("2025-01-01", "1-12月同比增长", values.iloc[4], "annual_cumulative")):
            yield _input(row, metric_id="industrial_production_growth_pct", value=value, unit="percent", currency=None, period_start=start, period_end="2025-12-31", fiscal_year=2025, fiscal_quarter=None, source_field_name=f"规模以上工业增加值:{field}", frequency=frequency, period_scope="calendar_period")
    elif record_key == "nbs_retail_sales_2025_12":
        values = pd.read_html(str(path))[0].iloc[2]
        for start, label, level, growth, frequency in (("2025-12-01", "12月", values.iloc[1], values.iloc[2], "monthly"), ("2025-01-01", "1-12月", values.iloc[3], values.iloc[4], "annual_cumulative")):
            yield _input(row, metric_id="retail_sales_current_cny", value=level, unit="hundred_million CNY", currency="CNY", period_start=start, period_end="2025-12-31", fiscal_year=2025, fiscal_quarter=None, source_field_name=f"社会消费品零售总额:{label}:绝对量", frequency=frequency, period_scope="calendar_period")
            yield _input(row, metric_id="retail_sales_growth_pct", value=growth, unit="percent", currency=None, period_start=start, period_end="2025-12-31", fiscal_year=2025, fiscal_quarter=None, source_field_name=f"社会消费品零售总额:{label}:同比增长", frequency=frequency, period_scope="calendar_period")
    elif record_key == "nbs_industrial_profits_2025":
        values = pd.read_html(str(path))[0].iloc[2]
        for metric_id, column, source_field in (("industrial_enterprise_revenue_current_cny", 1, "规模以上工业企业:营业收入"), ("industrial_enterprise_profit_current_cny", 5, "规模以上工业企业:利润总额")):
            yield _input(row, metric_id=metric_id, value=values.iloc[column], unit="hundred_million CNY", currency="CNY", period_start="2025-01-01", period_end="2025-12-31", fiscal_year=2025, fiscal_quarter="FY", source_field_name=source_field, frequency="annual", period_scope="calendar_period")
    elif record_key == "nbs_cpi_2025_12":
        text = " ".join(BeautifulSoup(path.read_bytes(), "lxml").get_text(" ").split())
        match = re.search(r"全国居民消费价格同比上涨\s*([0-9.]+)%", text)
        if match:
            yield _input(row, metric_id="inflation_rate_cpi", value=match.group(1), unit="percent", currency=None, period_start="2025-12-01", period_end="2025-12-31", fiscal_year=2025, fiscal_quarter=None, source_field_name="全国居民消费价格:12月同比", frequency="monthly", period_scope="calendar_period")
        annual = Decimal("0") if "2025 年全年，全国居民消费价格与上年持平" in text else None
        if annual is not None:
            yield _input(row, metric_id="inflation_rate_cpi", value=annual, unit="percent", currency=None, period_start="2025-01-01", period_end="2025-12-31", fiscal_year=2025, fiscal_quarter="FY", source_field_name="全国居民消费价格:全年涨跌幅", frequency="annual", period_scope="calendar_period")


def _pboc_inputs(row: dict[str, Any], path: Path) -> Iterable[dict[str, Any]]:
    record_key = str(row.get("record_key") or "")
    if path.suffix.lower() != ".html":
        return
    text = " ".join(BeautifulSoup(path.read_bytes(), "lxml").get_text(" ").split())
    if record_key == "pboc_financial_statistics_2024_07":
        match = re.search(r"广义货币\(M2\)余额([0-9.]+)万亿元,?同比增长([0-9.]+)%", text)
        if match:
            common = dict(period_start="2024-07-31", period_end="2024-07-31", fiscal_year=2024, fiscal_quarter=None, frequency="monthly", period_scope="calendar_point_in_time")
            yield _input(row, metric_id="money_supply_m2", value=match.group(1), unit="trillion CNY", currency="CNY", source_field_name="广义货币(M2)余额", **common)
            yield _input(row, metric_id="money_supply_m2_growth_pct", value=match.group(2), unit="percent", currency=None, source_field_name="广义货币(M2)同比增长", **common)
    elif record_key == "pboc_afre_flow_release_2025_08":
        match = re.search(r"前八个月社会融资规模增量累计为([0-9.]+)万亿元", text)
        if match:
            yield _input(row, metric_id="aggregate_financing_flow_current_cny", value=match.group(1), unit="trillion CNY", currency="CNY", period_start="2025-01-01", period_end="2025-08-31", fiscal_year=2025, fiscal_quarter=None, source_field_name="前八个月社会融资规模增量累计", frequency="monthly_cumulative", period_scope="calendar_period")


EXCHANGE_MARKET_FIELDS = {
    "listed_company_count": "No. of Listed Companies",
    "listed_security_count": "No. of Listed Securities",
    "market_capitalization": "Total Market Capitalization",
    "negotiable_market_capitalization": "Total Negotiable Market Capitalization",
    "market_turnover_value": "Total Turnover in Value",
    "market_average_pe_ratio": "Average P/E Ratio",
}


def _sse_inputs(row: dict[str, Any], path: Path) -> Iterable[dict[str, Any]]:
    if str(row.get("record_key") or "") != "sse_monthly_statistics_2021_12":
        return
    import fitz

    with fitz.open(path) as document:
        if len(document) <= 6:
            raise ValueError("SSE monthly report has no market-overview page 7")
        text = " ".join(document[6].get_text().split())

    specs = (
        ("listed_company_count", r"Number of Listed Companies\s+([0-9,.]+)", "count", None, "calendar_point_in_time"),
        ("listed_security_count", r"Number of Listed Securities\s+([0-9,.]+)", "count", None, "calendar_point_in_time"),
        ("market_capitalization", r"Total Market Capitalization \(100 Million RMB Yuan\)\s+([0-9,.]+)", "hundred_million CNY", "CNY", "calendar_point_in_time"),
        ("negotiable_market_capitalization", r"Total Market Capitalization Negotiable \(100 Million RMB Yuan\)\s+([0-9,.]+)", "hundred_million CNY", "CNY", "calendar_point_in_time"),
        ("market_turnover_value", r"Total Turnover in Value \(100 Million RMB Yuan\)\s+([0-9,.]+)", "hundred_million CNY", "CNY", "calendar_period"),
        ("market_average_pe_ratio", r"Weighted Average P/E Ratio\s+([0-9,.]+)", "ratio", None, "calendar_point_in_time"),
    )
    parsed = [
        (metric_id, _strict_labeled_decimal(text, pattern), unit, currency, period_scope)
        for metric_id, pattern, unit, currency, period_scope in specs
    ]
    for metric_id, value, unit, currency, period_scope in parsed:
        yield _input(
            row,
            metric_id=metric_id,
            value=value,
            unit=unit,
            currency=currency,
            period_start="2021-12-01" if period_scope == "calendar_period" else "2021-12-31",
            period_end="2021-12-31",
            fiscal_year=2021,
            fiscal_quarter=None,
            source_field_name=EXCHANGE_MARKET_FIELDS[metric_id],
            frequency="monthly",
            period_scope=period_scope,
            sheet="Market Overview / page 7",
        )


def _szse_inputs(row: dict[str, Any], path: Path) -> Iterable[dict[str, Any]]:
    if str(row.get("record_key") or "") != "szse_monthly_statistics_2025_08":
        return
    text = " ".join(BeautifulSoup(path.read_bytes(), "lxml").get_text(" ").split())
    specs = (
        ("listed_company_count", r"No\. of Listed Companies\s+([0-9,.]+)", "count", None, "calendar_point_in_time"),
        ("listed_security_count", r"No\. of Listed Securities\s+([0-9,.]+)", "count", None, "calendar_point_in_time"),
        ("market_capitalization", r"Total Market Capitalization \(RMB Mil\.\)\s+([0-9,.]+)", "million CNY", "CNY", "calendar_point_in_time"),
        ("negotiable_market_capitalization", r"Total Negotiable Market Capitalization \(RMB Mil\.\)\s+([0-9,.]+)", "million CNY", "CNY", "calendar_point_in_time"),
        ("market_turnover_value", r"Total Turnover \(RMB Mil\.\)\s+([0-9,.]+)", "million CNY", "CNY", "calendar_period"),
        ("market_average_pe_ratio", r"Average P/E Ratio at End of Month \(Times\)\s+([0-9,.]+)", "ratio", None, "calendar_point_in_time"),
    )
    parsed = [
        (metric_id, _strict_labeled_decimal(text, pattern), unit, currency, period_scope)
        for metric_id, pattern, unit, currency, period_scope in specs
    ]
    for metric_id, value, unit, currency, period_scope in parsed:
        yield _input(
            row,
            metric_id=metric_id,
            value=value,
            unit=unit,
            currency=currency,
            period_start="2025-08-01" if period_scope == "calendar_period" else "2025-08-31",
            period_end="2025-08-31",
            fiscal_year=2025,
            fiscal_quarter=None,
            source_field_name=EXCHANGE_MARKET_FIELDS[metric_id],
            frequency="monthly",
            period_scope=period_scope,
            sheet="Market Overview",
        )


def _strict_labeled_decimal(text: str, pattern: str) -> Decimal:
    match = re.search(pattern, text)
    if not match:
        raise ValueError(f"required exchange statistic label not found: {pattern}")
    value = _decimal(match.group(1))
    if value is None:
        raise ValueError(f"invalid exchange statistic value for label: {pattern}")
    return value


def _input(row: dict[str, Any], *, metric_id: str, value: Any, unit: str, currency: str | None, period_start: str, period_end: str, fiscal_year: int, fiscal_quarter: str | None, source_field_name: str, frequency: str, period_scope: str, sheet: str | None = None) -> dict[str, Any]:
    decimal_value = _decimal(value)
    if decimal_value is None:
        raise ValueError(f"non-numeric official value for {source_field_name}: {value!r}")
    return {
        "entity_code": row.get("entity_hint") or "CHN",
        "entity_name": "China",
        "metric_id": metric_id,
        "value": decimal_value,
        "unit": unit,
        "currency": currency,
        "period_start": period_start,
        "period_end": period_end,
        "fiscal_year": fiscal_year,
        "fiscal_quarter": fiscal_quarter,
        "as_of_date": period_end if "point_in_time" in period_scope else None,
        "report_date": row.get("source_publish_date") or period_end,
        "source_id": row.get("source_id"),
        "raw_object_id": row.get("raw_object_id"),
        "source_field_name": source_field_name,
        "extraction_method": "official_structured_publication",
        "confidence_score": 0.99,
        "verification_status": "single_source",
        "notes": {
            "record_key": row.get("record_key"),
            "frequency": frequency,
            "period_scope": period_scope,
            "sheet": sheet,
            "retrieval_time": str(row.get("retrieval_time") or ""),
            "authority_extraction": "deterministic_table_or_release_rule",
        },
        "stable_parts": [row.get("source_id"), row.get("record_key"), metric_id, source_field_name, period_start, period_end, str(decimal_value), unit],
    }


def _json_value(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    return {}


def _decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _bop_period(value: Any) -> tuple[str, str, int, str | None] | None:
    text = str(value or "").strip()
    if re.fullmatch(r"\d{4}", text):
        year = int(text)
        return f"{year}-01-01", f"{year}-12-31", year, "FY"
    match = re.fullmatch(r"(\d{4})Q([1-4])", text)
    if not match:
        return None
    year, quarter = int(match.group(1)), int(match.group(2))
    starts = {1: "01-01", 2: "04-01", 3: "07-01", 4: "10-01"}
    ends = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}
    return f"{year}-{starts[quarter]}", f"{year}-{ends[quarter]}", year, f"Q{quarter}"


def _cn_quarter_end(value: Any) -> tuple[str, str, int, str] | None:
    match = re.search(r"(\d{4})年([369]|12)月末", str(value or ""))
    if not match:
        return None
    year, month = int(match.group(1)), int(match.group(2))
    quarter = month // 3
    starts = {1: "01-01", 2: "04-01", 3: "07-01", 4: "10-01"}
    ends = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}
    return f"{year}-{starts[quarter]}", f"{year}-{ends[quarter]}", year, f"Q{quarter}"


def _month_period(value: Any) -> tuple[str, str, int] | None:
    if isinstance(value, datetime):
        year, month = value.year, value.month
    else:
        match = re.search(r"(\d{4})[.-](\d{1,2})", str(value or ""))
        if not match:
            return None
        year, month = int(match.group(1)), int(match.group(2))
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    end = (next_month - pd.Timedelta(days=1)).date().isoformat()
    return f"{year:04d}-{month:02d}-01", end, year
